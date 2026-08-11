from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from utils.assertion_executor import AssertionExecutor
from utils.case_validator import CaseValidationError, validate_cases
from utils.excel_handler import ExcelHandler
from utils.step_executor import StepExecutionError, StepExecutor


class WaitPage:
    def __init__(self):
        self.waited = []

    def wait_for_timeout(self, milliseconds):
        self.waited.append(milliseconds)


def test_fractional_wait_uses_milliseconds():
    page = WaitPage()
    StepExecutor(page).execute("0.5", "wait", "")
    assert page.waited == [500]


def test_unknown_operation_fails_instead_of_passing():
    with pytest.raises(StepExecutionError, match="不支持的操作类型"):
        StepExecutor(None).execute("", "clik", "")


class _InputPage:
    """记录 input 步骤最终填入的内容"""

    def __init__(self):
        self.filled = None

    def locator(self, _selector):
        return self

    @property
    def first(self):
        return self

    def wait_for(self, **kwargs):
        return None

    def fill(self, text):
        self.filled = text

    def wait_for_timeout(self, _ms):
        pass


def test_retry_report_locator_uses_pipe_separator():
    """retry_report 的定位器用 | 分隔三个部分，逗号分割后必须保持为一个定位器项"""
    locator = "text=查看报告|text=完 成|.ant-checkbox-group .ant-image"
    locs = [item.strip() for item in locator.split(",")]
    assert len(locs) == 1
    parts = [p.strip() for p in locs[0].split("|")]
    assert len(parts) == 3
    assert parts[0] == "text=查看报告" and parts[1] == "text=完 成"


def test_find_click_validation_locator_stays_single_item():
    """find_click 的「共检测|验证定位器」格式在逗号分割后必须保持为一个定位器项"""
    locs = [item.strip() for item in "共检测|.ant-image + div .anticon".split(",")]
    assert locs == ["共检测|.ant-image + div .anticon"]


def test_space_only_input_is_not_stripped_to_empty():
    """纯空格输入应保留为空格，而不是被 strip 成空串（TC-DETAIL-019 空格标签）"""
    page = _InputPage()
    StepExecutor(page).execute("input[placeholder='请输入标签内容']", "input", " ")
    assert page.filled == " "


def test_empty_input_still_fills_empty_string():
    page = _InputPage()
    StepExecutor(page).execute("input[placeholder='请输入标签内容']", "input", "")
    assert page.filled == ""


def test_date_range_parser_supports_slash_and_iso_dates():
    assert StepExecutor._parse_date_range("2026/1/2-2026/6/28") == (
        "2026-01-02",
        "2026-06-28",
    )
    assert StepExecutor._parse_date_range("2026-01-02~2026-06-28") == (
        "2026-01-02",
        "2026-06-28",
    )


def test_unknown_assertion_fails_instead_of_passing():
    with pytest.raises(AssertionError, match="未知断言类型"):
        AssertionExecutor(None).assert_by_type("visible_tex", "任意文本")


def test_visible_text_alias_uses_real_text_assertion():
    executor = AssertionExecutor(None)
    called = []
    executor._text_visible = lambda expected: called.append(expected) or True
    assert executor.assert_by_type("visible_text", "请输入账号") is True
    assert called == ["请输入账号"]


def test_text_hidden_dispatches_to_hidden_assertion():
    executor = AssertionExecutor(None)
    called = []
    executor._text_hidden = lambda expected: called.append(expected) or True
    assert executor.assert_by_type("text_hidden", "不显示'自动化标签测试1'标签") is True
    assert called == ["不显示'自动化标签测试1'标签"]


def test_element_disabled_dispatches_to_disabled_assertion():
    executor = AssertionExecutor(None)
    called = []
    executor._element_disabled = lambda locator: called.append(locator) or True
    assert executor.assert_by_type("element_disabled", "提交按钮禁用", "button.ant-btn-primary") is True
    assert called == ["button.ant-btn-primary"]


def test_value_equals_dispatches_to_value_assertion():
    executor = AssertionExecutor(None)
    called = []
    executor._value_equals = lambda expected, locator: called.append((expected, locator)) or True
    assert executor.assert_by_type("value_equals", "空", "input[placeholder='请输入家庭住址']") is True
    assert called == [("空", "input[placeholder='请输入家庭住址']")]


def _valid_case(**overrides):
    case = {
        "用例ID": "TC-DEMO-001",
        "操作类型": "wait",
        "元素定位器": "0.5",
        "输入数据": "",
        "断言类型": "url_contains",
        "验证点": "'/home'",
        "超时(秒)": "5",
        "_row": 2,
    }
    case.update(overrides)
    return case


def test_case_validator_rejects_invalid_wait_and_assertion():
    with pytest.raises(CaseValidationError) as exc_info:
        validate_cases([_valid_case(**{"元素定位器": "later", "断言类型": "unknown"})])
    message = str(exc_info.value)
    assert "wait 定位器必须是非负秒数" in message
    assert "不支持的断言类型" in message


def test_excel_reader_keeps_real_rows_and_batch_writes(tmp_path: Path):
    path = tmp_path / "cases.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "自动化测试用例"
    sheet.append(["用例ID", "模块", "实际结果"])
    sheet.append(["TC-A", "登录", ""])
    sheet.append([None, None, None])
    sheet.append(["TC-B", "首页", ""])
    workbook.save(path)
    workbook.close()

    handler = ExcelHandler(str(path))
    cases = handler.read_test_cases()
    assert [case["_row"] for case in cases] == [2, 4]

    handler.write_results([("TC-A", "pass", 2), ("TC-B", "fail: demo", 4)])
    workbook = load_workbook(path, read_only=True)
    sheet = workbook.active
    assert sheet.cell(2, 3).value == "pass"
    assert sheet.cell(4, 3).value == "fail: demo"
    # 失败用例应标红（浅红背景），通过用例不标红
    assert sheet.cell(4, 3).fill.start_color.rgb == "FFFFC7CE"
    assert sheet.cell(2, 3).fill.start_color.rgb != "FFFFC7CE"
    workbook.close()