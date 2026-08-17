"""
Excel 工具类 —— 负责从 Excel 读取测试用例，以及将测试结果回写。

支持两种格式：
  格式A（新版）: 用例ID、模块、测试场景、测试点、优先级、前置条件、
                  操作步骤、元素定位器、操作类型、输入数据、数据类型、
                  期望结果、验证点、断言类型、超时(秒)、备注、实际结果、
                  是否执行、执行分组
  格式B（旧版10列）: 编号、模块、功能-测试目的、功能点、前置条件、
                     输入动作、输入数据、期望结果、实际结果、备注
"""
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


class ExcelHandler:
    """处理测试用例 Excel 文件的读取与结果回写"""

    # 新旧格式的列名标识
    NEW_FORMAT_ID_COL = "用例ID"      # 新版用"用例ID"
    OLD_FORMAT_ID_COL = "编号"        # 旧版用"编号"

    def __init__(self, file_path: str):
        self.file_path = file_path
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"找不到用例文件: {file_path}")
        self._format = None  # "new" 或 "old"

    def detect_format(self) -> str:
        """自动检测 Excel 格式"""
        if self._format:
            return self._format
        wb = load_workbook(self.file_path, read_only=True)
        ws = wb.active
        first_row = [
            str(value or "").strip()
            for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        ]
        wb.close()
        if "用例ID" in first_row:
            self._format = "new"
        elif "编号" in first_row:
            self._format = "old"
        else:
            self._format = "unknown"
        return self._format

    def read_test_cases(self, sheet_name: str | int = 0) -> list[dict]:
        """
        读取所有测试用例，自动适配新旧格式
        :param sheet_name: 新版用 Sheet 名，旧版用索引
        """
        fmt = self.detect_format()
        if fmt == "new":
            return self._read_new_format()
        elif fmt == "old":
            return self._read_old_format(sheet_name)
        else:
            raise ValueError(f"无法识别 Excel 格式: {self.file_path}")

    def _read_new_format(self) -> list[dict]:
        """读取新版格式，并保留过滤前的真实 Excel 行号。"""
        df = pd.read_excel(self.file_path, sheet_name="自动化测试用例", engine="openpyxl")
        df.columns = df.columns.str.strip()
        df["_excel_row"] = df.index + 2

        if "模块" in df.columns:
            df["模块"] = df["模块"].ffill()
        if "用例ID" in df.columns:
            df = df.dropna(subset=["用例ID"])

        records = df.fillna("").to_dict(orient="records")
        for case in records:
            case["_row"] = int(case.pop("_excel_row"))
        return records

    def _read_old_format(self, sheet_name: str | int = 0) -> list[dict]:
        """读取旧版格式，并保留过滤前的真实 Excel 行号。"""
        df = pd.read_excel(self.file_path, sheet_name=sheet_name, engine="openpyxl")
        df.columns = df.columns.str.strip()
        df["_excel_row"] = df.index + 2

        if "编号" in df.columns:
            df["编号"] = df["编号"].ffill()
        if "模块" in df.columns:
            df["模块"] = df["模块"].ffill()
        if "功能 - 测试目的" in df.columns:
            df["功能 - 测试目的"] = df["功能 - 测试目的"].ffill()

        df = df.dropna(subset=["编号"])
        records = df.fillna("").to_dict(orient="records")
        for case in records:
            case["_row"] = int(case.pop("_excel_row"))
        return records
    def write_result(self, case_id: str, result: str, row_num: int | None = None) -> None:
        """写入单条结果；内部复用批量实现。"""
        self.write_results([(case_id, result, row_num)])

    def write_results(self, results: list[tuple[str, str, int | None]]) -> None:
        """一次打开并保存 Excel，批量写入多条测试结果。"""
        if not results:
            return

        wb = load_workbook(self.file_path)
        try:
            if self.detect_format() == "new" and "自动化测试用例" in wb.sheetnames:
                ws = wb["自动化测试用例"]
            else:
                ws = wb.active

            headers = {
                str(ws.cell(1, column).value).strip(): column
                for column in range(1, ws.max_column + 1)
                if ws.cell(1, column).value
            }
            id_column_name = "用例ID" if "用例ID" in headers else "编号" if "编号" in headers else None
            if not id_column_name:
                raise ValueError("Excel 找不到用例ID列")
            id_column = headers[id_column_name]

            result_column = headers.get("实际结果")
            if not result_column:
                result_column = ws.max_column + 1
                ws.cell(1, result_column).value = "实际结果"

            row_by_id = {
                str(ws.cell(row, id_column).value).strip(): row
                for row in range(2, ws.max_row + 1)
                if ws.cell(row, id_column).value
            }
            missing: list[str] = []
            # 失败标红（浅红背景+深红加粗文字），通过/跳过重置样式，避免旧红色残留
            red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
            red_font = Font(color="FF9C0006", bold=True)
            for case_id, result, row_num in results:
                target_row = int(row_num) if row_num else row_by_id.get(str(case_id).strip())
                if not target_row or target_row < 2 or target_row > ws.max_row:
                    missing.append(str(case_id))
                    continue
                cell = ws.cell(target_row, result_column)
                cell.value = result
                result_text = str(result).lower()
                if result_text.startswith("fail") or "error" in result_text:
                    cell.fill = red_fill
                    cell.font = red_font
                else:
                    cell.fill = PatternFill()
                    cell.font = Font()

            if missing:
                raise ValueError(f"以下用例未找到对应 Excel 行: {', '.join(missing)}")
            wb.save(self.file_path)
        finally:
            wb.close()
